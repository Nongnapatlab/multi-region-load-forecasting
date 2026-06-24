"""Combine all pipeline outputs into a single multi-sheet Excel workbook
for easy review — predictions, metrics, daily MAPE, diagnostics, and the
decision dashboard all in one file."""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
BODY_FONT = Font(name="Arial")


def _write_sheet(writer, df: pd.DataFrame, sheet_name: str) -> None:
    if df is None or df.empty:
        df = pd.DataFrame({"note": ["No data for this run"]})
    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def _format_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    ws.freeze_panes = "A2"


def build_excel_report(
    output_path: Path,
    dashboard_df: pd.DataFrame,
    predictions_df: pd.DataFrame,
    metrics_df: pd.DataFrame,
    daily_mape_df: pd.DataFrame,
    diagnostics_df: pd.DataFrame,
    bias_summary_df: pd.DataFrame = None,
    today_df: pd.DataFrame = None,
    future_df: pd.DataFrame = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets = [
        ("dashboard", dashboard_df),
        ("predictions_all", predictions_df),
        ("metrics", metrics_df),
        ("daily_mape", daily_mape_df),
        ("diagnostics", diagnostics_df),
    ]
    if bias_summary_df is not None:
        sheets.append(("bias_summary", bias_summary_df))
    if today_df is not None:
        sheets.append(("today", today_df))
    if future_df is not None:
        sheets.append(("future", future_df))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            _write_sheet(writer, df, sheet_name)
        for sheet_name, _ in sheets:
            _format_sheet(writer.sheets[sheet_name[:31]])

    return output_path